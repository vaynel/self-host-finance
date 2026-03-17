"""Smart parser for CSV/Excel files with flexible column mapping."""

import csv
import io
import re
from typing import Any, Optional
from datetime import datetime


def _normalize_header_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    return s.replace(" ", "").replace("_", "").replace("-", "")


def _header_score(cells: list[Any]) -> int:
    """
    Heuristic score for whether a row looks like a header row.
    Higher is more header-like.
    """
    header_keywords = [
        "날짜",
        "date",
        "거래일",
        "거래일자",
        "일자",
        "내역",
        "거래내역",
        "적요",
        "내용",
        "금액",
        "amount",
        "입금",
        "출금",
        "수입",
        "지출",
        "잔액",
        "balance",
        "구분",
        "type",
        "카테고리",
        "category",
        "계좌",
        "account",
        "메모",
        "memo",
        "비고",
    ]

    normalized = [_normalize_header_cell(c) for c in cells]
    if not normalized or all(x == "" for x in normalized):
        return 0

    hits = 0
    for cell in normalized:
        for kw in header_keywords:
            if kw in cell:
                hits += 1
                break

    non_numeric = 0
    for cell in normalized:
        if cell == "":
            continue
        if re.fullmatch(r"[0-9.,()/+-]+", cell):
            continue
        non_numeric += 1

    if hits < 2:
        return hits
    return hits * 10 + non_numeric


def _find_header_row_index(rows: list[list[Any]]) -> int:
    """Find header row index within scanned rows. Defaults to 0 if unsure."""
    best_i = 0
    best_score = 0
    for i, r in enumerate(rows):
        score = _header_score(r)
        if score > best_score:
            best_score = score
            best_i = i
    return best_i


def extract_csv_headers_and_samples(content: bytes, encoding: str = "utf-8-sig", sample_n: int = 5) -> tuple[list[str], list[list[Any]]]:
    """Extract CSV headers and a few raw sample rows (as lists)."""
    text = content.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    scan_rows: list[list[Any]] = []
    for _ in range(50):
        r = next(reader, None)
        if r is None:
            break
        scan_rows.append(r)
    if not scan_rows:
        return [], []

    header_i = _find_header_row_index(scan_rows)
    headers = scan_rows[header_i]

    samples: list[list[Any]] = []
    for r in scan_rows[header_i + 1 :]:
        if len(samples) >= sample_n:
            break
        samples.append(r)
    while len(samples) < sample_n:
        r = next(reader, None)
        if r is None:
            break
        samples.append(r)

    return [str(h).strip() for h in headers], samples


def extract_xlsx_headers_and_samples(content: bytes, sample_n: int = 5) -> tuple[list[str], list[list[Any]]]:
    """Extract XLSX headers and a few raw sample rows (as lists) from the first sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    scan_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if row is None:
            continue
        scan_rows.append(list(row))
    if not scan_rows:
        wb.close()
        return [], []

    header_i = _find_header_row_index(scan_rows)
    headers = [str(v).strip() if v is not None else "" for v in scan_rows[header_i]]

    samples: list[list[Any]] = []
    for r in scan_rows[header_i + 1 :]:
        if len(samples) >= sample_n:
            break
        samples.append(r)
    wb.close()
    return headers, samples


def extract_xls_headers_and_samples(content: bytes, sample_n: int = 5) -> tuple[list[str], list[list[Any]]]:
    """Extract XLS headers and a few raw sample rows (as lists) from the first sheet."""
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)

    scan_rows: list[list[Any]] = []
    max_r = min(sheet.nrows, 50)
    for r in range(0, max_r):
        scan_rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    if not scan_rows:
        return [], []

    header_i = _find_header_row_index(scan_rows)
    headers = [str(v).strip() if v is not None else "" for v in scan_rows[header_i]]

    samples: list[list[Any]] = []
    for r in scan_rows[header_i + 1 :]:
        if len(samples) >= sample_n:
            break
        samples.append(r)
    # continue if needed
    rr = header_i + 1
    while len(samples) < sample_n and rr + 1 < sheet.nrows:
        rr += 1
        samples.append([sheet.cell_value(rr, c) for c in range(sheet.ncols)])
    return headers, samples


def detect_column_mapping(headers: list[str]) -> dict[str, Optional[int]]:
    """Detect column mapping from headers for various bank formats."""
    mapping: dict[str, Optional[int]] = {
        "date": None,
        "description": None,
        "amount": None,
        "type": None,
        "category": None,
        "account": None,
        "memo": None,
        "balance": None,  # 잔액 (optional)
    }
    
    # Normalize headers (lowercase, strip, remove special chars)
    normalized_headers = []
    for i, h in enumerate(headers):
        if h:
            normalized = str(h).strip().lower().replace(" ", "").replace("_", "").replace("-", "")
            normalized_headers.append((i, normalized, h))
    
    # Date column detection
    date_patterns = ["날짜", "date", "거래일자", "거래일", "일자", "일시", "거래시각"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in date_patterns):
            mapping["date"] = idx
            break
    
    # Description column detection
    desc_patterns = ["내역", "거래내역", "description", "적요", "거래적요", "내용", "거래내용", "메모", "비고", "적요명"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in desc_patterns):
            mapping["description"] = idx
            break
    
    # Amount column detection
    amount_patterns = ["금액", "amount", "거래금액", "입금", "출금", "수입", "지출", "변동금액"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in amount_patterns):
            mapping["amount"] = idx
            break
    
    # Balance column detection (optional)
    balance_patterns = ["잔액", "balance", "거래후잔액", "잔고"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in balance_patterns):
            mapping["balance"] = idx
            break
    
    # Type column detection (optional)
    type_patterns = ["구분", "type", "유형", "거래구분", "입출금"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in type_patterns):
            mapping["type"] = idx
            break
    
    # Category column detection (optional)
    category_patterns = ["카테고리", "category", "분류", "항목"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in category_patterns):
            mapping["category"] = idx
            break
    
    # Account column detection (optional)
    account_patterns = ["계좌", "account", "계좌번호", "계좌명", "은행"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in account_patterns):
            mapping["account"] = idx
            break
    
    # Memo column detection (optional)
    memo_patterns = ["메모", "memo", "비고", "참고", "기타"]
    for idx, norm, orig in normalized_headers:
        if any(pattern in norm for pattern in memo_patterns):
            mapping["memo"] = idx
            break
    
    return mapping


def apply_index_mapping(base: dict[str, Optional[int]], override: dict[str, Optional[int]] | None) -> dict[str, Optional[int]]:
    """
    Merge an override index mapping into the base mapping.
    """
    if not override:
        return base
    merged = dict(base)
    for k, v in override.items():
        if k in merged and v is not None:
            merged[k] = v
    return merged


def parse_amount(value: Any) -> float:
    """Parse amount from various formats."""
    if value is None:
        return 0.0
    
    # Convert to string
    str_value = str(value).strip()
    
    # Remove currency symbols and spaces
    str_value = re.sub(r'[₩$€£,\s]', '', str_value)
    
    # Handle negative amounts (parentheses, minus sign)
    is_negative = False
    if str_value.startswith('(') and str_value.endswith(')'):
        is_negative = True
        str_value = str_value[1:-1]
    elif str_value.startswith('-'):
        is_negative = True
        str_value = str_value[1:]
    
    try:
        amount = float(str_value)
        return -amount if is_negative else amount
    except (ValueError, TypeError):
        return 0.0


def parse_date(value: Any) -> Optional[str]:
    """Parse date from various formats."""
    if value is None:
        return None
    
    # If it's already a datetime object
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    
    str_value = str(value).strip()
    
    # Try common date formats
    date_formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y.%m.%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y%m%d',
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(str_value, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    # If all formats fail, try to extract YYYY-MM-DD pattern
    match = re.search(r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', str_value)
    if match:
        year, month, day = match.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    
    return None


def parse_csv_smart(
    content: bytes,
    encoding: str = "utf-8-sig",
    *,
    index_mapping_override: dict[str, Optional[int]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Optional[int]]]:
    """Parse CSV with smart column detection. Returns (rows, column_mapping)."""
    text = content.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    
    # Scan initial rows to find header row (header may not be first)
    scan_rows: list[list[Any]] = []
    for _ in range(50):
        r = next(reader, None)
        if r is None:
            break
        scan_rows.append(r)
    if not scan_rows:
        return [], {}
    
    header_i = _find_header_row_index(scan_rows)
    headers = [str(h).strip() for h in scan_rows[header_i]]

    # Detect column mapping
    mapping = apply_index_mapping(detect_column_mapping(headers), index_mapping_override)
    
    rows = []
    # process remaining rows from scan first
    for row in scan_rows[header_i + 1 :]:
        if not row or all(not cell or str(cell).strip() == "" for cell in row):
            continue
        
        date_val = parse_date(row[mapping["date"]]) if mapping["date"] is not None and mapping["date"] < len(row) else None
        description = str(row[mapping["description"]]).strip() if mapping["description"] is not None and mapping["description"] < len(row) else ""
        amount_val = parse_amount(row[mapping["amount"]]) if mapping.get("amount") is not None and mapping["amount"] < len(row) else 0.0
        type_val = str(row[mapping["type"]]).strip().lower() if mapping["type"] is not None and mapping["type"] < len(row) else None
        category_val = str(row[mapping["category"]]).strip() if mapping["category"] is not None and mapping["category"] < len(row) else None
        account_val = str(row[mapping["account"]]).strip() if mapping["account"] is not None and mapping["account"] < len(row) else None
        memo_val = str(row[mapping["memo"]]).strip() if mapping["memo"] is not None and mapping["memo"] < len(row) else None
        
        if not date_val or not description:
            continue
        
        rows.append({
            "date": date_val,
            "description": description,
            "amount": amount_val,
            "type": type_val,
            "category": category_val,
            "account": account_val,
            "memo": memo_val or "",
        })

    # then continue reading
    for row in reader:
        if not row or all(not cell or str(cell).strip() == "" for cell in row):
            continue
        
        # Extract values using mapping
        date_val = parse_date(row[mapping["date"]]) if mapping["date"] is not None and mapping["date"] < len(row) else None
        description = str(row[mapping["description"]]).strip() if mapping["description"] is not None and mapping["description"] < len(row) else ""
        amount_val = parse_amount(row[mapping["amount"]]) if mapping.get("amount") is not None and mapping["amount"] < len(row) else 0.0
        type_val = str(row[mapping["type"]]).strip().lower() if mapping["type"] is not None and mapping["type"] < len(row) else None
        category_val = str(row[mapping["category"]]).strip() if mapping["category"] is not None and mapping["category"] < len(row) else None
        account_val = str(row[mapping["account"]]).strip() if mapping["account"] is not None and mapping["account"] < len(row) else None
        memo_val = str(row[mapping["memo"]]).strip() if mapping["memo"] is not None and mapping["memo"] < len(row) else None
        
        # Skip if essential fields are missing
        if not date_val or not description:
            continue
        
        rows.append({
            "date": date_val,
            "description": description,
            "amount": amount_val,
            "type": type_val,
            "category": category_val,
            "account": account_val,
            "memo": memo_val or "",
        })
    
    return rows, mapping


def parse_xlsx_smart(
    content: bytes,
    *,
    index_mapping_override: dict[str, Optional[int]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Optional[int]]]:
    """Parse XLSX with smart column detection. Returns (rows, column_mapping)."""
    import openpyxl
    
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    # Scan initial rows to find header row (header may not be first)
    scan_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if row is None:
            continue
        scan_rows.append(list(row))
    if not scan_rows:
        wb.close()
        return [], {}
    
    header_i = _find_header_row_index(scan_rows)
    headers = [str(v).strip() if v is not None else "" for v in scan_rows[header_i]]

    # Detect column mapping
    mapping = apply_index_mapping(detect_column_mapping(headers), index_mapping_override)
    
    rows = []
    min_row = header_i + 2  # 1-based
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if not row or all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue
        
        # Extract values using mapping
        date_val = parse_date(row[mapping["date"]]) if mapping["date"] is not None and mapping["date"] < len(row) else None
        description = str(row[mapping["description"]]).strip() if mapping["description"] is not None and mapping["description"] < len(row) else ""
        amount_val = parse_amount(row[mapping["amount"]]) if mapping["amount"] is not None and mapping["amount"] < len(row) else 0.0
        type_val = str(row[mapping["type"]]).strip().lower() if mapping["type"] is not None and mapping["type"] < len(row) else None
        category_val = str(row[mapping["category"]]).strip() if mapping["category"] is not None and mapping["category"] < len(row) else None
        account_val = str(row[mapping["account"]]).strip() if mapping["account"] is not None and mapping["account"] < len(row) else None
        memo_val = str(row[mapping["memo"]]).strip() if mapping["memo"] is not None and mapping["memo"] < len(row) else None
        
        # Skip if essential fields are missing
        if not date_val or not description:
            continue
        
        rows.append({
            "date": date_val,
            "description": description,
            "amount": amount_val,
            "type": type_val,
            "category": category_val,
            "account": account_val,
            "memo": memo_val or "",
        })
    
    wb.close()
    return rows, mapping


def parse_xls_smart(
    content: bytes,
    *,
    index_mapping_override: dict[str, Optional[int]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Optional[int]]]:
    """Parse legacy XLS with smart column detection. Returns (rows, column_mapping)."""
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)

    if sheet.nrows < 1:
        return [], {}

    scan_rows: list[list[Any]] = []
    max_r = min(sheet.nrows, 50)
    for r in range(0, max_r):
        scan_rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    if not scan_rows:
        return [], {}

    header_i = _find_header_row_index(scan_rows)
    headers = [str(v).strip() if v is not None else "" for v in scan_rows[header_i]]

    mapping = apply_index_mapping(detect_column_mapping(headers), index_mapping_override)

    rows: list[dict[str, Any]] = []
    for r in range(header_i + 1, sheet.nrows):
        row_vals = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if not row_vals or all((v is None) or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            continue

        date_val = parse_date(row_vals[mapping["date"]]) if mapping["date"] is not None and mapping["date"] < len(row_vals) else None
        description = str(row_vals[mapping["description"]]).strip() if mapping["description"] is not None and mapping["description"] < len(row_vals) else ""
        amount_val = parse_amount(row_vals[mapping["amount"]]) if mapping["amount"] is not None and mapping["amount"] < len(row_vals) else 0.0
        type_val = str(row_vals[mapping["type"]]).strip().lower() if mapping["type"] is not None and mapping["type"] < len(row_vals) else None
        category_val = str(row_vals[mapping["category"]]).strip() if mapping["category"] is not None and mapping["category"] < len(row_vals) else None
        account_val = str(row_vals[mapping["account"]]).strip() if mapping["account"] is not None and mapping["account"] < len(row_vals) else None
        memo_val = str(row_vals[mapping["memo"]]).strip() if mapping["memo"] is not None and mapping["memo"] < len(row_vals) else None

        if not date_val or not description:
            continue

        rows.append(
            {
                "date": date_val,
                "description": description,
                "amount": amount_val,
                "type": type_val,
                "category": category_val,
                "account": account_val,
                "memo": memo_val or "",
            }
        )

    return rows, mapping
